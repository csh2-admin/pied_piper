"""
excel_export.py — Generate a formatted Excel report from database rows.
Called on-demand from the GUI; does NOT write to a shared file continuously.
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import EXCEL_OUTPUT_PATH

COLUMNS = [
    ("ID",                          "id",                        8),
    ("Timestamp (UTC)",             "logged_at",                 22),
    ("Engineer",                    "engineer",                  16),
    ("Source File",                 "source_file",               28),
    ("Activity Type",               "activity_type",             26),
    ("Summary",                     "summary",                   45),
    ("System Performance",          "system_performance",        45),
    ("Action Items",                "action_items",              45),
    ("Components Affected",         "components_affected",       30),
    ("Duration (hrs)",              "duration_hours",            14),
    ("Trigger Simulation Update",   "trigger_simulation_update", 14),
    ("Additional Notes",            "additional_notes",          45),
    ("Raw Transcript",              "raw_transcript",            60),
]

HEADER_FILL  = PatternFill("solid", start_color="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT    = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", start_color="EEF2F7")
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
THIN         = Side(style="thin", color="C0C8D8")
THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def export_to_excel(rows: list[dict], output_path: str = EXCEL_OUTPUT_PATH) -> str:
    """
    Write rows (from db_logger.fetch_all_rows) to a formatted .xlsx file.
    Returns the output path.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Maintenance Log"
    ws.freeze_panes = "A2"

    # Header
    ws.row_dimensions[1].height = 30
    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_num, row in enumerate(rows, start=2):
        is_alt = (row_num % 2 == 0)
        ws.row_dimensions[row_num].height = 55

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            value = row.get(key, "")
            # Format datetime nicely
            if hasattr(value, "strftime"):
                value = value.strftime("%Y-%m-%d %H:%M:%S UTC")
            # Render booleans as Y/N for the simulation flag
            if key == "trigger_simulation_update":
                value = "Y" if bool(value) else "N"

            cell = ws.cell(row=row_num, column=col_idx, value=str(value) if value is not None else "")
            cell.font      = DATA_FONT
            cell.alignment = WRAP_ALIGN
            cell.border    = THIN_BORDER

            if is_alt:
                cell.fill = ALT_FILL

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Export generated"
    ws2["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws2["A2"] = "Total entries"
    ws2["B2"] = len(rows)
    ws2["A3"] = "Engineers"
    engineers = sorted({r.get("engineer", "") for r in rows if r.get("engineer")})
    ws2["B3"] = ", ".join(engineers)

    act_counts = {}
    for r in rows:
        a = r.get("activity_type", "Unknown") or "Unknown"
        act_counts[a] = act_counts.get(a, 0) + 1
    ws2["A5"] = "Activity type breakdown"
    for i, (act, count) in enumerate(sorted(act_counts.items()), start=6):
        ws2[f"A{i}"] = act
        ws2[f"B{i}"] = count

    sim_count = sum(1 for r in rows if r.get("trigger_simulation_update"))
    ws2[f"A{6 + len(act_counts) + 1}"] = "Entries flagged for simulation update"
    ws2[f"B{6 + len(act_counts) + 1}"] = sim_count

    for col in ["A", "B"]:
        ws2.column_dimensions[col].width = 24

    wb.save(output_path)
    return output_path
